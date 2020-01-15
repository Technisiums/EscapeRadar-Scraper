from functools import partial
from PyQt5 import QtCore, QtGui, QtWidgets
from design import Ui_MainWindow
from openpyxl import load_workbook
import sys
from Phase1 import *
from Phase2 import *


class ButtonData:
    def __init__(self):
        self.id = ''
        self.url = ''
        self.avg = ''
        self.room_name = ''
        self.city = ''
        self.country = ''


class mywindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(mywindow, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.ui.lineEdit.setPlaceholderText("Search By Room Name, City or Country")
        scroll_area = self.ui.GridScrolarea
        scrollAreaWidgetContents = QtWidgets.QWidget()
        scroll_area.setWidgetResizable(True)
        self.grid = QtWidgets.QGridLayout(scrollAreaWidgetContents)
        scroll_area.setWidget(scrollAreaWidgetContents)
        self.ui.verticalLayout_2.addWidget(scroll_area)
        self.data = list()
        self.read_excel_file('input/input.xlsx')
        self.add_items(self.data)
        self.ui.pushButton.clicked.connect(self.search_click)
        self.ui.pushButton_2.clicked.connect(partial(self.add_items, self.data))
        self.ui.pushButton_3.clicked.connect(partial(self.run_combo_click))

    def run_combo_click(self):
        t = self.ui.comboBox.currentText()
        print(t)
        koi_b_phase2(t)

    def read_excel_file(self, file):
        wb = load_workbook(file)
        s = wb[wb.sheetnames[0]]
        max = s.max_row
        for row in range(2, max + 1):
            obj = ButtonData()
            obj.id = s['A' + str(row)].value
            obj.url = s['B' + str(row)].value
            obj.avg = s['C' + str(row)].value
            obj.room_name = s['D' + str(row)].value
            obj.city = s['E' + str(row)].value
            obj.country = s['F' + str(row)].value
            self.data.append(obj)
        wb.close()

    def clear_all(self):
        for i in range(self.grid.count()):
            self.grid.itemAt(i).widget().close()

    def add_items(self, data):
        self.clear_all()
        rows = int(len(data) / 5) + 1
        count = 0
        for i in range(rows):
            for j in range(5):
                if count == len(data):
                    break
                btn = QtWidgets.QPushButton(data[count].room_name)
                btn.setFixedHeight(30)
                btn.setFixedWidth(150)
                btn.clicked.connect(partial(self.button_clicked, data[count]))
                self.grid.addWidget(btn, i, j)
                count = count + 1

    def button_clicked(self, data):
        print(data.id)
        koi_b(data)

    def search_click(self):
        self.clear_all()
        key = str(self.ui.lineEdit.text()).lower().strip()
        subData = list()
        for d in self.data:
            if key in str(d.city).lower().strip() or key in str(d.room_name).lower().strip() or key in str(
                    d.country).lower().strip():
                subData.append(d)
        if len(subData) == 0:
            lbl = QtWidgets.QLabel("No Result Found")
            font = QtGui.QFont()
            font.setPointSize(20)
            lbl.setFont(font)
            self.grid.addWidget(lbl, 0, 0)
        else:
            self.add_items(subData)


# _append_run_path()
print("Tool Started")
app = QtWidgets.QApplication([])
application = mywindow()
application.show()
sys.exit(app.exec())
