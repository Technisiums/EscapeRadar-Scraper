from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
import datetime
from datetime import datetime as dt
from openpyxl import Workbook
from openpyxl import load_workbook
from design import Ui_MainWindow
from PyQt5 import QtCore, QtGui, QtWidgets


class ExcelWriter:
    workbook = None
    worksheet = None
    count = 1
    fname = ""

    def __init__(self, filename):
        self.fname = "output/" + filename + ".xlsx"
        try:
            self.workbook = load_workbook(self.fname)
            self.worksheet = self.workbook[self.workbook.sheetnames[0]]
            print("Excel Already Exists")
        except:
            self.workbook = Workbook()
            self.workbook.save(self.fname)
            self.worksheet = self.workbook[self.workbook.sheetnames[0]]
            self.write_headers(1)
            print("Creating a New Excel Sheet")
        finally:
            self.workbook.save(self.fname)

    def write_headers(self, row):
        row = str(row)
        self.worksheet['A' + row].value = 'Escape Room Name'
        self.worksheet['B' + row].value = 'URL'
        self.worksheet['C' + row].value = 'Room/Service'
        self.worksheet['D' + row].value = 'Region'
        self.worksheet['E' + row].value = 'City'
        self.worksheet['F' + row].value = 'Country'
        self.worksheet['G' + row].value = 'Average Booking Price'
        self.worksheet['H' + row].value = 'Scraping Year'
        self.worksheet['I' + row].value = 'Scraping Month'
        self.worksheet['J' + row].value = 'Scraping Day of the Month'
        self.worksheet['K' + row].value = 'Scraping Week of the Year'
        self.worksheet['L' + row].value = 'Scraping Day of the Week Number'
        self.worksheet['M' + row].value = 'Scraping Hour of the Day'
        self.worksheet['N' + row].value = 'Booking Year'
        self.worksheet['O' + row].value = 'Booking Month'
        self.worksheet['P' + row].value = 'Booking Day of the Month'
        self.worksheet['Q' + row].value = 'Booking Week of the Year'
        self.worksheet['R' + row].value = 'Booking Day of the Week Number'
        self.worksheet['S' + row].value = 'Booking Hour of the Day'
        self.worksheet['T' + row].value = 'Booked'
        self.worksheet['U' + row].value = 'Booking Value'
        self.worksheet['V' + row].value = 'Escape Radar Rating'
        self.worksheet['W' + row].value = 'Region Ranking'
        self.worksheet['X' + row].value = 'Country Ranking'
        self.worksheet['Y' + row].value = 'Booking Feature'
        self.workbook.save(self.fname)

    def write(self, games):
        print("Writing Data to Excel")
        row = str(self.worksheet.max_row + 1)
        for game in games:
            for seat in game.seats:
                self.worksheet['A' + row].value = game.room_name
                self.worksheet['B' + row].value = game.link
                self.worksheet['C' + row].value = game.game_name
                self.worksheet['D' + row].value = game.city
                self.worksheet['E' + row].value = game.main_city
                self.worksheet['F' + row].value = game.country
                self.worksheet['G' + row].value = float(game.avg)
                ds = game.date_scraper.split(',')
                self.worksheet['H' + row].value = int(ds[0])
                self.worksheet['I' + row].value = int(ds[1])
                self.worksheet['J' + row].value = int(ds[2])
                self.worksheet['K' + row].value = int(ds[3])
                self.worksheet['L' + row].value = int(ds[4])
                self.worksheet['M' + row].value = int(ds[5])
                db = seat.date_booking.split(',')
                self.worksheet['N' + row].value = int(db[0])
                self.worksheet['O' + row].value = int(db[1])
                self.worksheet['P' + row].value = int(db[2])
                self.worksheet['Q' + row].value = int(db[3])
                self.worksheet['R' + row].value = int(db[4])
                self.worksheet['S' + row].value = int(db[5])
                self.worksheet['T' + row].value = int(seat.status)
                if seat.status == '1':
                    self.worksheet['U' + row].value = float(game.avg)
                else:
                    self.worksheet['U' + row].value = int('0')
                self.worksheet['V' + row].value = game.radar_ranking
                if game.region_ranking != '':
                    self.worksheet['W' + row].value = int(game.region_ranking)
                if game.country_ranking != '':
                    self.worksheet['X' + row].value = int(game.country_ranking)
                self.worksheet['Y' + row].value = int(game.booking_feature)
                row = str(int(row) + 1)
        self.workbook.save(self.fname)
        print("Data has been Written")


class Seats:
    def __init__(self):
        self.date_booking = ''
        self.status = ''


class GameData:
    def __init__(self):
        self.room_name = ''
        self.date_scraper = ''
        self.game_name = ''
        self.avg = ''
        self.link = ''
        self.seats = list()
        self.city = ''
        self.main_city = ''
        self.country = ''
        self.country_ranking = ''
        self.region_ranking = ''
        self.radar_ranking = ''
        self.booking_feature = ''


class Scraper:
    def __init__(self, city):
        print("Initializing Browser")
        self.browser = Chrome()
        self.games = list()
        self.prices = dict()
        self.writer = ExcelWriter(city)

    def calculate_week(self, year, month, date):
        return datetime.date(int(year), int(month), int(date)).isocalendar()[1]

    def calculate_weekday(self, year, month, date):
        dt = datetime.datetime.strptime(str(year) + '-' + str(month) + '-' + str(date), "%Y-%m-%d")
        return dt.weekday() + 1

    def get_current_date_time(self):
        # year;month;date;week;day;hour
        year = str(dt.now().year)
        month = str(dt.now().month)
        date = str(dt.now().day)
        hour = str(dt.now().hour)
        week = str(datetime.date(int(year), int(month), int(date)).isocalendar()[1])
        day = str(datetime.date(int(year), int(month), int(date)).weekday() + 1)
        dd = year + ',' + month + ',' + date + ',' + week + ',' + day + ',' + hour
        return dd

    def calculate_booking_date(self, y, m, d):
        week = self.calculate_week(y, m, d)
        weekday = self.calculate_weekday(y, m, d)
        return str(y) + ',' + str(m) + ',' + str(d) + ',' + str(week) + ',' + str(weekday) + ','

    def show_more(self):
        results = self.browser.find_elements(By.CSS_SELECTOR, 'div.no-gutters.item')
        len_of_result = len(results)
        while True:
            print("Scrolling/Clicking on Show more")
            self.browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            try:
                self.browser.find_element(By.ID, 'btn_cookie').click()
            except:
                pass
            sleep(1.5)
            try:
                btn = WebDriverWait(self.browser, 5).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, 'a.showMore-btn')))
                btn.click()
            except:
                print("no more games to show")
                break

    def scrape(self, url):
        self.browser.get(url)
        for x in range(9):
            print("Moving to Day: ", x + 1)
            d = self.browser.find_elements(By.CSS_SELECTOR, 'button.btn-link.btn-date')[1].get_attribute('data-value')
            self.browser.get(url + '--' + str(d))
            print("Scrolling")
            self.show_more()
            results = self.browser.find_elements(By.CSS_SELECTOR, 'div.no-gutters.item')
            print(len(results), "Games to Scrap")
            self.games = list()
            region_rankings = self.browser.find_elements(By.XPATH, '//span[@title="Ranking comunidad autónoma"]')
            radar_rankings = self.browser.find_elements(By.XPATH, '//span[@title="Puntuación"]')
            country_ranking = self.browser.find_elements(By.XPATH, '//span[@title="Ranking país"]')
            cities_names = self.browser.find_elements(By.XPATH, '//strong//following-sibling::a')
            date = str(d).split('/')[0]
            month = str(d).split('/')[1]
            year = str(d).split('/')[2]
            print("Scraping for", date, '/', month, '/', year)
            # year;month;date;week;day;hour
            bdate = self.calculate_booking_date(year, month, date)
            currdate = self.get_current_date_time()
            for i in range(len(results)):
                print("=============================================\nScraping ", i + 1, 'out of ', len(results), 'Day',
                      x + 1)
                g = GameData()
                g.date_scraper = currdate
                g.link = url
                # region
                try:
                    k = str(region_rankings[i].text).strip().split('/')[0].strip()
                    if k != '-':
                        g.radar_ranking = k
                except:
                    pass
                try:
                    k = str(country_ranking[i].text).strip().split('/')[0].strip()
                    if k != '-':
                        g.country_ranking = k
                except:
                    pass
                try:
                    results[i].find_element(By.CSS_SELECTOR,
                                            'span.btn.btn-ver-todas-horas.btn-dark.btn-sm.py-0.text-white ').click()
                except:
                    pass
                try:
                    k = str(radar_rankings[i].text).strip()
                    if k != '-':
                        g.radar_ranking = k
                except:
                    pass
                g.main_city = \
                    str(self.browser.find_element(By.CSS_SELECTOR, 'h2.text-capitalize').text).strip().split(',')[0]
                g.city = str(cities_names[i].text).strip()
                g.room_name = str(
                    results[i].find_element(By.TAG_NAME, 'strong').find_element(By.TAG_NAME, 'a').text).strip()
                g.game_name = str(results[i].find_element(By.CSS_SELECTOR, 'h5.mb-0').text).strip()
                blocks = results[i].find_elements(By.CSS_SELECTOR, 'li.list-inline-item.hour-block')
                for hour in blocks:
                    seat = Seats()
                    h = str(hour.get_attribute('innerHTML')).strip().split('</span>')[1]
                    seat.date_booking = bdate + str(h).strip().split(':')[0]
                    # print(hour.text)
                    if 'hour-disabled' in str(hour.get_attribute('class')):
                        seat.status = '1'
                    else:
                        seat.status = '0'
                    g.seats.append(seat)
                temp = results[i].find_element(By.CSS_SELECTOR, 'div.col-auto.ml-auto.d-none.d-lg-block')
                try:
                    temp.find_element(By.TAG_NAME, 'button')
                    g.booking_feature = '1'
                except:
                    g.booking_feature = '0'
                self.games.append(g)
                # print('\nEscape name', g.room_name, '\nGame Name', g.game_name, '\nRegion Rating', g.region_ranking,
                #       '\nradar', g.radar_ranking,
                #       '\nCountry: ',
                #       g.country_ranking, "\nCity: ", g.city, '\nBooking Feature: ', g.booking_feature)
            self.prewriter()
            self.writer.write(self.games)

    def prewriter(self):
        for game in self.games:
            try:
                val = self.prices[game.room_name.lower()]
            except:
                val = self.prices['average']
            game.avg = val

    def excel_reader(self, city):
        wb = load_workbook('input/input1.xlsx', data_only=True)
        sheet = wb[city]
        self.prices['link'] = str(sheet['C1'].value)
        for x in range(2, sheet.max_row + 1):
            key = str(sheet['B' + str(x)].value).strip().lower()
            val = str(sheet['C' + str(x)].value).strip()
            self.prices[key] = val
        wb.close()
        print(self.prices)

    def RUN(self, city):
        self.excel_reader(city)
        self.scrape(self.prices['link'])


def koi_b_phase2(city):
    obj = Scraper(city)
    obj.RUN(city)
