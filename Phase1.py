from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
import datetime
from datetime import datetime as dt
from openpyxl import Workbook
from openpyxl import load_workbook


class Seats:
    def __init__(self):
        self.date_booking = ''
        self.status = ''


class GameData:
    def __init__(self):
        self.room_name = ''
        self.date_scraper = ''
        self.game_name = ''
        self.avg = ''
        self.link = ''
        self.seats = list()
        self.city = ''
        self.country = ''


class Scraper:
    def __init__(self):
        print("Initializing Browser")
        self.browser = Chrome()
        self.writer = ExcelWriter()

    def cal_month(self, month):
        c = ''
        month = str(month).lower()
        if month == 'genn' or month == 'gennaio' or month == 'enero':
            c = 1
        if month == 'febbr' or month == 'febbraio' or month == 'febrero':
            c = 2
        if month == 'mar' or month == 'marzo' or month == 'marzo':
            c = 3
        if month == 'apr' or month == 'aprile' or month == 'abril':
            c = 4
        if month == 'magg' or month == 'maggio' or month == 'mayo':
            c = 5
        if month == 'giugno' or month == 'giugno' or month == 'junio':
            c = 6
        if month == 'luglio' or month == 'luglio' or month == 'julio':
            c = 7
        if month == 'ag' or month == 'agosto' or month == 'agosto':
            c = 8
        if month == 'sett' or month == 'settembre' or month == 'septiembre':
            c = 9
        if month == 'ott' or month == 'ottobre' or month == 'octubre':
            c = 10
        if month == 'nov' or month == 'novembre' or month == 'noviembre':
            c = 11
        if month == 'dic' or month == 'dicembre' or month == 'diciembre':
            c = 12
        return c

    def calculate_week(self, year, month, date):
        return datetime.date(int(year), int(month), int(date)).isocalendar()[1]

    def calculate_weekday(self, year, month, date):
        dt = datetime.datetime.strptime(str(year) + '-' + str(month) + '-' + str(date), "%Y-%m-%d")
        return dt.weekday() + 1

    def get_current_date_time(self):
        # year;month;date;week;day;hour
        year = str(dt.now().year)
        month = str(dt.now().month)
        date = str(dt.now().day)
        hour = str(dt.now().hour)
        week = str(datetime.date(int(year), int(month), int(date)).isocalendar()[1])
        day = str(datetime.date(int(year), int(month), int(date)).weekday() + 1)
        dd = year + ',' + month + ',' + date + ',' + week + ',' + day + ',' + hour
        return dd

    def calculate_booking_date(self, y, m, d):
        week = self.calculate_week(y, m, d)
        weekday = self.calculate_weekday(y, m, d)
        return str(y) + ',' + str(m) + ',' + str(d) + ',' + str(week) + ',' + str(weekday) + ','

    def scrape_Fox_in_a_Box_Madrid(self, link, avg, room_name, city, country):
        print("=========================================================")
        print("Scraping Started for:", room_name)
        self.browser.get(link)
        # GET BOOKING DATE
        for k in range(0, 15):
            print("_____________________________________________________")
            print("Day", k + 1)
            try:
                WebDriverWait(self.browser, 60).until(
                    EC.invisibility_of_element_located(
                        (By.CSS_SELECTOR, 'div.bookingTableContainerWorking')))
            except:
                print("Something went wrong")
                continue
            calender = self.browser.find_element(By.CSS_SELECTOR, 'div.col-md-6.col-sm-12.text-center.fox-color')
            d = str(calender.find_element(By.CSS_SELECTOR, 'span.cp').text)
            date = str(d.split(' ')[2])
            # print(d.split(' ')[1],' Month: ',month)
            calender.find_element(By.CSS_SELECTOR, 'span.glyphicon.glyphicon-calendar').click()
            y = str(
                self.browser.find_element(By.CSS_SELECTOR, 'table.table.table-condensed').find_element(By.CSS_SELECTOR,
                                                                                                       'th.datepicker-switch').text)
            month = y.split(' ')[0]
            year = y.split(' ')[1]
            month = str(self.cal_month(month))

            table = self.browser.find_element(By.ID, 'multipleBookingTable')
            bdate = self.calculate_booking_date(year, month, date)
            # END GET BOOKING DATE
            names = table.find_elements(By.CSS_SELECTOR, 'th.day_heading.day_heading_madrid')
            games = list()
            print("Scraping Started for: ", date, ':', month, ':', year)
            for i in range(len(names) - 1):
                n = names[i]
                try:
                    name = str(n.find_elements(By.TAG_NAME, 'font')[1].get_attribute('innerHTML'))
                except:
                    name = str(n.find_element(By.TAG_NAME, 'b').get_attribute('innerHTML'))
                finally:
                    obj = GameData()
                    for g in games:
                        if name == g.game_name:
                            name = name + ' 2'
                    obj.game_name = name
                    obj.avg = avg
                    obj.link = link
                    obj.room_name = room_name
                    obj.city = city
                    obj.country = country
                    obj.date_scraper = self.get_current_date_time()
                    games.append(obj)

            table = table.find_element(By.TAG_NAME, 'tbody')
            trs = table.find_elements(By.TAG_NAME, 'tr')

            for j in range(len(trs)):
                tr = trs[j]
                tds = tr.find_elements(By.TAG_NAME, 'td')
                for i in range(len(tds)):
                    td = tds[i]
                    if i == 0:
                        hour = str(td.find_element(By.TAG_NAME, 'b').get_attribute('innerHTML'))
                        continue
                    if j == 0 and i == len(tds) - 1:
                        continue
                    status = str(
                        td.find_element(By.CSS_SELECTOR, 'span.visible-lg').get_attribute('innerHTML')).strip().lower()
                    seat = Seats()
                    if status == '&nbsp;':
                        continue
                    if status == 'reservado':
                        seat.status = '1'
                    if status == 'elige':
                        seat.status = '0'
                    seat.date_booking = bdate + str(hour).split(':')[0]
                    games[i - 1].seats.append(seat)
            print("Scraping Finished for: ", date, ':', month, ':', year)
            self.writer.write_row(games)
            try:
                self.browser.find_element(By.CSS_SELECTOR, 'a.cc-btn.cc-dismiss').click()
                sleep(2)
            except:
                pass
            print("Moving to Next Date")
            js = str(self.browser.find_elements(By.CSS_SELECTOR, 'span.glyphicon.glyphicon-chevron-right.fox-color.cp')[
                         0].get_attribute('onclick'))
            self.browser.execute_script(js)

    def scraper_The_Rombo_Code_Madrid(self, link, avg, room_name, city, country):
        print("=========================================================")
        print("Scraping Started for:", room_name)
        self.browser.get(link)
        for k in range(15):
            print("_____________________________________________________")
            print("Day", k + 1)
            try:
                WebDriverWait(self.browser, 60).until(
                    EC.invisibility_of_element_located(
                        (By.CSS_SELECTOR, 'div.preloader')))
            except:
                print("Something Went Wrong")
                continue
            games = list()
            d = str(self.browser.find_element(By.CSS_SELECTOR, 'div.filter-item-medium').find_element(By.TAG_NAME,
                                                                                                      'input').get_attribute(
                'value'))
            date = d.split('-')[0]
            month = d.split('-')[1]
            year = d.split('-')[2]
            b = self.calculate_booking_date(year, month, date)
            main_table = self.browser.find_elements(By.CSS_SELECTOR, 'table.booking-daily-table.center')[
                0].find_element(
                By.TAG_NAME, 'tbody')
            trs = main_table.find_elements(By.TAG_NAME, 'tr')
            print("Scraping Started for: ", date, ':', month, ':', year)
            for tr in trs:
                game = GameData()
                seat = Seats()
                hour = str(tr.find_element(By.CSS_SELECTOR, 'td.booking-daily-col1').text).split(':')[0]
                name = \
                    str(tr.find_element(By.CSS_SELECTOR, 'td.booking-daily-col2').get_attribute('innerHTML')).split(
                        '<div')[
                        0]
                status = tr.find_element(By.CSS_SELECTOR, 'td.booking-daily-col5')
                try:
                    status.find_element(By.CSS_SELECTOR, 'div.booking-daily-not-available')
                    seat.status = '1'
                except:
                    status.find_element(By.CSS_SELECTOR, 'a.booking-daily-buy-now')
                    seat.status = '0'
                seat.date_booking = self.calculate_booking_date(year, month, date) + hour.split(':')[0].strip()

                game.game_name = name
                game.date_scraper = self.get_current_date_time()
                game.link = link
                game.avg = avg
                game.room_name = room_name
                game.city = city
                game.country = country
                flag = False
                for g in games:
                    if name == g.game_name:
                        g.seats.append(seat)
                        flag = True
                if not flag:
                    game.seats.append(seat)
                    games.append(game)

            print("Scraping Finished for: ", date, ':', month, ':', year)
            self.writer.write_row(games)
            print("Moving to Next Date")
            self.browser.find_elements(By.CSS_SELECTOR, 'div.filter-item-tiny')[1].click()

    def scraper_Coco_Room_Madrid(self, link, avg, room_name, city, country):
        print("=========================================================")
        print("Scraping Started for:", room_name)
        self.browser.get(link)
        sleep(1)
        select = self.browser.find_element(By.CSS_SELECTOR,
                                           'select.form-control.ng-pristine.ng-untouched.ng-valid')
        select.send_keys('4')
        sleep(2)
        # select.form-control.ng-pristine.ng-untouched.ng-valid
        for k in range(15):
            print("_____________________________________________________")
            print("Day", k + 1)
            calbtn = self.browser.find_element(By.CSS_SELECTOR, 'input.form-control.ng-valid-date')
            self.browser.execute_script('window.scroll(0,800)')
            calbtn.click()
            btn = self.browser.find_element(By.CSS_SELECTOR, 'button.btn.btn-default.btn-sm.active')
            if k != 0:
                btnid = str(btn.find_element(By.XPATH, '..').get_attribute('id')).strip()
                fid = btnid.split('-')
                lid = int(fid[-1]) + 1
                id = ''
                for x in range(len(fid) - 1):
                    id = id + str(fid[x]) + '-'
                id = id + str(lid)
                btn = self.browser.find_element(By.ID, id).find_element(By.TAG_NAME, 'button')
            date = btn.find_element(By.TAG_NAME, 'span').get_attribute('innerHTML')
            btn.click()
            sleep(2)
            self.browser.execute_script('window.scroll(0,800)')
            calbtn = self.browser.find_element(By.CSS_SELECTOR, 'input.form-control.ng-valid-date')
            calbtn.click()
            y = str(self.browser.find_element(By.CSS_SELECTOR, 'strong.ng-binding').text)
            year = y.split(' ')[1]
            month = self.cal_month(y.split(' ')[0])
            sleep(1)
            b_date = self.calculate_booking_date(year, month, date)
            # add try catch
            boxes = self.browser.find_elements(By.XPATH,
                                               '//div[@ng-repeat="act in activities"]//div[@class="ng-scope"]')
            games = list()
            print("Scraping Started for: ", date, ':', month, ':', year)
            for g in boxes:
                obj = GameData()
                name = str(g.find_element(By.CSS_SELECTOR, 'h4.ng-binding').text)
                obj.game_name = name
                obj.room_name = room_name
                obj.link = link
                obj.avg = avg
                obj.city = city
                obj.country = country
                obj.date_scraper = self.get_current_date_time()
                lis = g.find_element(By.TAG_NAME, 'ul').find_elements(By.TAG_NAME, 'li')
                for li in lis:
                    seat = Seats()
                    c = li.find_element(By.XPATH, '//span[@ng-show="validation_date"]').get_attribute('class')
                    if 'ng-hide' in c:
                        seat.status = '1'
                    else:
                        seat.status = '0'
                    seat.date_booking = b_date + str(li.text).strip().split(':')[0]
                    obj.seats.append(seat)
                games.append(obj)
            print("Scraping Finished for: ", date, ':', month, ':', year)
            self.writer.write_row(games)
            print("Moving to Next Date")

    def RUN(self, obj):
        if obj.id == 1:
            self.scrape_Fox_in_a_Box_Madrid(obj.url, obj.avg, obj.room_name, obj.city, obj.country)
        if obj.id == 2:
            self.scraper_The_Rombo_Code_Madrid(obj.url, obj.avg, obj.room_name, obj.city, obj.country)
        if obj.id == 3:
            self.scraper_Coco_Room_Madrid(obj.url, obj.avg, obj.room_name, obj.city, obj.country)


def koi_b(o):
    obj = Scraper()
    obj.RUN(o)
    print("Scrapping has been Finished for", o.room_name)
    print("Closing Browser")
    try:
        obj.browser.close()
        obj.browser.quit()
    except:
        pass


class ExcelWriter:
    workbook = None
    worksheet = None
    count = 1
    fname = ""

    def __init__(self):
        self.fname = "output/outfile.xlsx"
        try:
            self.workbook = load_workbook(self.fname)
            self.worksheet = self.workbook[self.workbook.sheetnames[0]]
            print("Excel Already Exists")
        except:
            self.workbook = Workbook()
            self.workbook.save(self.fname)
            self.worksheet = self.workbook[self.workbook.sheetnames[0]]
            self.write_headers(1)
            print("Creating a New Excel Sheet")
        finally:
            self.workbook.save(self.fname)

    def write_headers(self, row):
        row = str(row)
        self.worksheet['A' + row].value = 'Escape Room Name'
        self.worksheet['B' + row].value = 'URL'
        self.worksheet['C' + row].value = 'Room/Service'
        self.worksheet['D' + row].value = 'City'
        self.worksheet['E' + row].value = 'Country'
        self.worksheet['F' + row].value = 'Average Booking Price'
        self.worksheet['G' + row].value = 'Scraping Year'
        self.worksheet['H' + row].value = 'Scraping Month'
        self.worksheet['I' + row].value = 'Scraping Day of the Month'
        self.worksheet['J' + row].value = 'Scraping Week of the Year'
        self.worksheet['K' + row].value = 'Scraping Day of the Week Number'
        self.worksheet['L' + row].value = 'Scraping Hour of the Day'
        self.worksheet['M' + row].value = 'Booking Year'
        self.worksheet['N' + row].value = 'Booking Month'
        self.worksheet['O' + row].value = 'Booking Day of the Month'
        self.worksheet['P' + row].value = 'Booking Week of the Year'
        self.worksheet['Q' + row].value = 'Booking Day of the Week Number'
        self.worksheet['R' + row].value = 'Booking Hour of the Day'
        self.worksheet['S' + row].value = 'Booked'
        self.worksheet['T' + row].value = 'Booking Value'
        self.workbook.save(self.fname)

    def write_row(self, games):
        print("Writing Data to Excel")
        row = str(self.worksheet.max_row + 1)
        for game in games:
            for seat in game.seats:
                self.worksheet['A' + row].value = game.room_name
                self.worksheet['B' + row].value = game.link
                self.worksheet['C' + row].value = game.game_name
                self.worksheet['D' + row].value = game.city
                self.worksheet['E' + row].value = game.country
                self.worksheet['F' + row].value = int(game.avg)
                ds = game.date_scraper.split(',')
                self.worksheet['G' + row].value = int(ds[0])
                self.worksheet['H' + row].value = int(ds[1])
                self.worksheet['I' + row].value = int(ds[2])
                self.worksheet['J' + row].value = int(ds[3])
                self.worksheet['K' + row].value = int(ds[4])
                self.worksheet['L' + row].value = int(ds[5])
                db = seat.date_booking.split(',')
                self.worksheet['M' + row].value = int(db[0])
                self.worksheet['N' + row].value = int(db[1])
                self.worksheet['O' + row].value = int(db[2])
                self.worksheet['P' + row].value = int(db[3])
                self.worksheet['Q' + row].value = int(db[4])
                self.worksheet['R' + row].value = int(db[5])
                self.worksheet['S' + row].value = int(seat.status)
                if seat.status == '1':
                    self.worksheet['T' + row].value = int(game.avg)
                else:
                    self.worksheet['T' + row].value = int('0')
                row = str(int(row) + 1)
        self.workbook.save(self.fname)
        print("Data has been Written")
